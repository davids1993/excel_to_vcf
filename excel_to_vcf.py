# from interface import missing_headers

#get excel file location
def get_excel_file():
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    return filedialog.askopenfilename()

#check file type matches xlsx,.xlsm,.xltx,.xltm
def check_file_type(file_location):
    import pathlib
    file_location = pathlib.Path(file_location)
    file_type = file_location.suffix
    if file_type == '.xlsx' or file_type == '.xlsm' or file_type == '.xltx' or file_type == '.xltm':
        return True
    else:
        import sys
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror('Error', f'File type {file_type} is not supported')
        sys.exit()


#get the first row of an excel spreadsheet as a dictionary key with column letter as value
def get_headers(sheet):
    headers = {}
    for column in range(1, sheet.max_column + 1):
        headers[sheet.cell(row=1, column=column).value] = sheet.cell(row=1, column=column).column_letter
    return headers

#check if name, email, phone, and title are in the excel spreadsheet headers
def check_headers(sheet):
    headers = get_headers(sheet)
    if 'name' in headers and 'phone' in headers:
        return True
    else:
        import sys
        import tkinter as tk
        from tkinter import messagebox
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror('Error', 'One of these headers is missing: name, phone')
        sys.exit()

#import excel contact list and map columns to headers
def import_contacts(file_location):
    import openpyxl
    wb = openpyxl.load_workbook(file_location)
    sheet = wb.worksheets[0]
    headers = get_headers(sheet)
    check_headers(sheet)
    contacts = []
    for row in range(2, sheet.max_row + 1):
        contact = {}
        contact['name'] = sheet[headers['name'] + str(row)].value
        contact['phone'] = sheet[headers['phone'] + str(row)].value
        if 'email' in headers:
            contact['email'] = sheet[headers['email'] + str(row)].value
        if 'title' in headers:
            contact['title'] = sheet[headers['title'] + str(row)].value
        if 'phone2' in headers:
            contact['phone2'] = sheet[headers['phone2'] + str(row)].value
        contacts.append(contact)
    return contacts

#remove characters that are not allowed in vcf files
def remove_chars(string):
    import re
    string = re.sub(r'[^\w]', ' ', string)
    return string

#convert contacts to .vcf
def contacts_to_vcf(contacts):
    import vobject
    vcfs = []
    for contact in contacts:
        vcard = vobject.vCard()
        # vcard.add('n')
        # vcard.n.value = vobject.vcard.Name(family=contact['name'])
        if contact['name']:
            vcard.add('fn')
            vcard.fn.value = remove_chars(contact['name'])
        if contact['phone']:
            vcard.add('tel')
            vcard.tel.value = remove_chars(contact['phone'])
            vcard.tel.type_param = 'CELL'
        if 'email' in contact and contact['email']:
            vcard.add('email')
            vcard.email.value = remove_chars(contact['email'])
            vcard.email.type_param = 'INTERNET'
        if 'phone2' in contact and contact['phone2']:
            vcard.add('tel2')
            vcard.tel2.value = remove_chars(contact['phone2'])
            vcard.tel2.type_param = 'WORK'
        if 'title' in contact and contact['title']:
            vcard.add('title')
            vcard.title.value = remove_chars(contact['title'])
        vcfs.append(vcard)
    return vcfs

#create new folder named 'contacts'
def create_folder(file_location):
    import os
    import shutil
    file_location = get_path(file_location)
    os.chdir(file_location)
    try:
        os.mkdir('contacts')
        os.mkdir('merged_contacts')
    except FileExistsError:
        #delete old folder and create new one
        shutil.rmtree('contacts')
        shutil.rmtree('merged_contacts')
        os.mkdir('contacts')
        os.mkdir('merged_contacts')


#go to file location and save .vcf files to all contacts folder
def save_vcfs(file_location, vcfs):
    import os
    file_location = get_path(file_location)
    os.chdir(file_location)
    os.chdir('contacts')
    for vcf in vcfs:
        try:
            with open(f'{vcf.fn.value}.vcf', 'w') as f:
                f.write(vcf.serialize())
        except:
            print(f"{vcf.fn.value} doesn't exists")

            
#get path without the file name
def get_path(file_location):
    import os
    return os.path.dirname(file_location)

#merge all text files into one
def merge_files(file_location):
    file_location = get_path(file_location)
    import os
    import glob
    os.chdir(file_location)
    os.chdir('contacts')
    files = glob.glob('*.vcf')
    save_path = os.path.join(file_location, 'merged_contacts')

    with open(os.path.join(save_path, 'contacts.vcf'), 'w') as outfile:
        for fname in files:
            with open(fname, 'r') as infile:
                outfile.write(infile.read())

                
